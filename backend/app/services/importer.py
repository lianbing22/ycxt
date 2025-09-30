"""Utilities to convert uploaded tabular files into structured records."""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Protocol

try:  # Optional dependency for Excel parsing
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - best effort optional import
    load_workbook = None  # type: ignore

from ..storage import UploadMetadata, UploadStorage


class AnalysisBackend(Protocol):
    """Protocol describing the analysis hook invoked after import."""

    def run(self, upload: UploadMetadata, records: List[Dict[str, object]]) -> Dict[str, object]:
        """Return analysis results for the provided records."""


@dataclass
class ImportJob:
    """Represents the lifecycle of an import execution."""

    id: str
    upload_id: str
    status: str
    started_at: datetime
    finished_at: datetime
    record_count: int
    analysis: Dict[str, object] = field(default_factory=dict)
    error: str | None = None

    def to_dict(self) -> Dict[str, object]:
        payload = {
            "id": self.id,
            "upload_id": self.upload_id,
            "status": self.status,
            "started_at": self.started_at.isoformat(),
            "finished_at": self.finished_at.isoformat(),
            "record_count": self.record_count,
            "analysis": self.analysis,
        }
        if self.error:
            payload["error"] = self.error
        return payload


class SimpleAnalysis(AnalysisBackend):
    """Fallback analysis backend that summarises the import."""

    def run(self, upload: UploadMetadata, records: List[Dict[str, object]]) -> Dict[str, object]:
        columns = sorted({key for row in records for key in row.keys()})
        return {
            "upload_id": upload.id,
            "total_rows": len(records),
            "columns": columns,
        }


class Importer:
    """Parse uploaded files and trigger downstream analysis."""

    def __init__(self, storage: UploadStorage, analysis_backend: AnalysisBackend | None = None) -> None:
        self._storage = storage
        self._analysis = analysis_backend or SimpleAnalysis()

    def enqueue(self, upload: UploadMetadata) -> ImportJob:
        start = datetime.utcnow()
        try:
            records = self._parse(Path(upload.stored_path))
            analysis = self._analysis.run(upload, records)
            status = "completed"
            error = None
        except Exception as exc:  # pragma: no cover - defensive
            records = []
            analysis = {}
            status = "failed"
            error = str(exc)

        finish = datetime.utcnow()
        upload.status = status
        upload.record_count = len(records)
        upload.error = error
        self._storage.update(upload)
        return ImportJob(
            id=f"job-{upload.id}",
            upload_id=upload.id,
            status=status,
            started_at=start,
            finished_at=finish,
            record_count=len(records),
            analysis=analysis,
            error=error,
        )

    def _parse(self, path: Path) -> List[Dict[str, object]]:
        extension = path.suffix.lower()
        if extension == ".csv":
            return self._parse_csv(path)
        if extension in {".xls", ".xlsx"}:
            return self._parse_excel(path)
        raise ValueError(f"Unsupported file type: {extension}")

    def _parse_csv(self, path: Path) -> List[Dict[str, object]]:
        with path.open("r", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            return [dict(row) for row in reader]

    def _parse_excel(self, path: Path) -> List[Dict[str, object]]:
        if load_workbook is None:
            raise RuntimeError("Excel support requires the 'openpyxl' package")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.rows)
        if not rows:
            return []
        headers = [str(cell.value) if cell.value is not None else "" for cell in rows[0]]
        records: List[Dict[str, object]] = []
        for row in rows[1:]:
            record = {header: cell.value for header, cell in zip(headers, row)}
            records.append(record)
        return records
